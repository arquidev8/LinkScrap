


import selenium.webdriver.support.expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
import time
import pandas as pd
import os
from openpyxl import load_workbook




import json
import xml.etree.ElementTree as ET
from selenium import webdriver
from selenium.common import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
import pandas as pd
from datetime import date
import time
import re
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains


# options = webdriver.ChromeOptions()
# options.add_experimental_option("detach", True)
# options.add_experimental_option("excludeSwitches", ["enable-automation"])
# options.add_experimental_option('useAutomationExtension', False)
# options.add_argument("--disable-blink-features=AutomationControlled")
# options.add_argument('--disable-extensions')
# options.add_argument('--no-sandbox')
# options.add_argument('--disable-infobars')
# options.add_argument('--disable-dev-shm-usage')
# options.add_argument('--disable-browser-side-navigation')
# options.add_argument('--disable-gpu')
#
# driver = webdriver.Chrome(options=options)
#
# # Define la URL base y el número total de páginas a scrappear
# base_url = "https://www.linkedin.com/search/results/people/?geoUrn=%5B%22105646813%22%5D&keywords=consultoria&origin=FACETED_SEARCH&profileLanguage=%5B%22es%22%5D&sid=mTd&titleFreeText=CEO"
# total_pages =  10  # Cambia esto al número total de páginas que deseas scrappear
#
# # Función para extraer las URLs de los perfiles
# def extract_profile_links():
#     profile_links = []
#     # Asume que siempre hay  10 elementos por página
#     xpath = "//span[@class='entity-result__title-line entity-result__title-line--2-lines ']/span/a"
#     try:
#         # Usa .find_elements para obtener una lista de todos los elementos que coinciden
#         elements = driver.find_elements(By.XPATH, xpath)
#         for element in elements:
#             # Para cada elemento, obtén el atributo href y agrégalo a la lista de enlaces
#             href = element.get_attribute("href")
#             profile_links.append(href)
#     except NoSuchElementException:
#         print("No se encontraron elementos.")
#         pass
#     return profile_links
#
#
#
# # Inicia sesión en LinkedIn
# driver.execute_script("window.open('https://www.linkedin.com/home', '_blank')")
# time.sleep(5)
# driver.switch_to.window(driver.window_handles[1])
#
# time.sleep(3)
# username_field = driver.find_element(By.XPATH, '//*[@id="session_key"]').send_keys("hectorcreatives08@gmail.com")
# time.sleep(3)
# password_field = driver.find_element(By.XPATH, '//*[@id="session_password"]').send_keys("20759364")
# time.sleep(3)
# btn = driver.find_element(By.XPATH, '//*[@id="main-content"]/section[1]/div/div/form/div[2]/button')
# btn.click()
# time.sleep(5)
#
# # Navega a la primera página de resultados de búsqueda
# driver.get(base_url + "&page=1")
# time.sleep(5)  # Espera a que la página se cargue completamente
#
# all_links = []
#
# # Itera a través de todas las páginas y extrae las URLs
# for page in range(1, total_pages +  1):
#     # Extrae las URLs de la página actual
#     profile_links = extract_profile_links()
#     # Agrega los enlaces a la lista de todos los enlaces
#     all_links.extend(profile_links)
#
#     # Imprime las URLs en la consola
#     for link in profile_links:
#         print(link)
#
#     # Navega a la siguiente página cambiando el parámetro 'page' en la URL
#     if page < total_pages:
#         next_page = page +  1
#         driver.get(base_url + f"&page={next_page}")
#         time.sleep(20)  # Espera a que la página siguiente se cargue completamente
#
# # Convertir la lista de enlaces en un DataFrame de pandas
# df = pd.DataFrame(all_links, columns=['Enlace'])
#
# # Nombre del archivo donde se guardarán las URLs
# output_file = "linkedin_links.xlsx"
#
# # Asegúrate de que el archivo exista antes de intentar abrirlo
# if not os.path.exists(output_file):
#     # Si el archivo no existe, crea uno nuevo con los encabezados
#     df.to_excel(output_file, index=False)
# else:
#     # Si el archivo ya existe, lee los datos existentes y concatena con los nuevos
#     existing_data = pd.read_excel(output_file)
#     combined_data = pd.concat([existing_data, df], ignore_index=True)
#     # Guarda los datos combinados en el archivo de Excel
#     combined_data.to_excel(output_file, index=False)
#
# # Cierra el navegador
# driver.quit()
#
#



#ULTIMA VERSION FUNCIONAL
# import pandas as pd
# import time
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.common.exceptions import TimeoutException
#
# # Configuración de las opciones del navegador
# options = webdriver.ChromeOptions()
# options.add_experimental_option("detach", True)
# options.add_experimental_option("excludeSwitches", ["enable-automation"])
# options.add_experimental_option('useAutomationExtension', False)
# options.add_argument("--disable-blink-features=AutomationControlled")
# options.add_argument('--disable-extensions')
# options.add_argument('--no-sandbox')
# options.add_argument('--disable-infobars')
# options.add_argument('--disable-dev-shm-usage')
# options.add_argument('--disable-browser-side-navigation')
# options.add_argument('--disable-gpu')
# # options.add_argument(r"C:\Users\hecto\AppData\Local\Google\Chrome\User Data\Persona 1")
#
# # options.page_load_strategy = 'none'
#
# # Iniciar el navegador con las opciones configuradas
# driver = webdriver.Chrome(options=options)
#
# # Iniciar sesión en LinkedIn
# driver.execute_script("window.open('https://www.linkedin.com/home')")
# time.sleep(10)
# driver.switch_to.window(driver.window_handles[1])
# time.sleep(12)
#
# username_field = driver.find_element(By.XPATH, '//*[@id="session_key"]').send_keys("arqhector08@gmail.com")
# time.sleep(3)
# password_field = driver.find_element(By.XPATH, '//*[@id="session_password"]').send_keys("alejandro20759364")
# time.sleep(5)
# btn = driver.find_element(By.XPATH, '//*[@id="main-content"]/section[1]/div/div/form/div[2]/button')
# btn.click()
# time.sleep(15)
#
#
# # Leer el archivo Excel y obtener los URLs de la columna "Referencia"
# df = pd.read_excel('linkedin_links.xlsx', sheet_name='Sheet1', usecols=['Enlace'])
#
# # Convertir los URLs en una lista
# url_list = df['Enlace'].tolist()
#
#
#
# # Definir un diccionario que mapee el número de intento al XPath correspondiente
# xpath_dict = {
#     1: "//section[@class='pv-contact-info__contact-type'][2]/ul/li/a",  # Web
#     2: "//*[@class='pv-contact-info__contact-type'][3]/ul/li/span",         # Teléfono
#     3: "//section[@class='pv-contact-info__contact-type'][4]/div/a"     # Email
# }
#
# # Recorrer cada URL de la lista
# for url in url_list:
#     max_attempts = 3  # Establecer el número total de intentos
#
#     for attempt in range(1, max_attempts + 1):
#         # try:
#         driver.get(url)
#         time.sleep(40)
#
#         # Extraer el nombre solo en la primera iteración
#         if attempt == 1:
#             try:
#                 name = driver.find_element(By.XPATH, "//div[@class='oSsIKjSDcyduLrRLQBBCYBpUDyeLCY']/span/a/h1")
#                 value_name = name.text
#             except NoSuchElementException:
#                 print(f"No se encontró el elemento en la URL: {url}")
#                 value_name = "N/A"
#
#         # Intenta presionar el botón de información
#         try:
#             information_button = WebDriverWait(driver, 30).until(
#                 EC.element_to_be_clickable((By.ID, 'top-card-text-details-contact-info'))
#             )
#             actions = ActionChains(driver)
#             actions.move_to_element(information_button).click().perform()
#
#         except NoSuchElementException:
#             print(f"No se encontró el botón de información para la URL: {url}")
#             continue
#
#         time.sleep(30)
#
#         # Intenta encontrar el elemento correspondiente al intento actual
#         xpath = xpath_dict.get(attempt)
#         if xpath:
#             try:
#                 element = driver.find_element(By.XPATH, xpath)
#                 value = element.text if attempt != 1 else element.get_attribute("href")
#             except NoSuchElementException:
#                 print(f"No se encontró el elemento para el intento {attempt} en la URL: {url}")
#                 value = "N/A"
#         else:
#             print(f"No se proporcionó un XPath para el intento {attempt}")
#
#         # Imprimir el valor del elemento
#         if attempt == 1:
#             print("\nDatos de Perfil:")
#         print(f"Elemento {attempt}: {value}, {value_name}")
#
# # Cerrar el navegador
# driver.close()
#
#




#SEGUNDA OPCCION
# import pandas as pd
# import time
# from selenium import webdriver
# from selenium.webdriver.common.by import By
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support import expected_conditions as EC
# from selenium.common.exceptions import TimeoutException
# import openpyxl
#
#
# # Configuración de las opciones del navegador
# options = webdriver.ChromeOptions()
# options.add_experimental_option("detach", True)
# options.add_experimental_option("excludeSwitches", ["enable-automation"])
# options.add_experimental_option('useAutomationExtension', False)
# options.add_argument("--disable-blink-features=AutomationControlled")
# options.add_argument('--disable-extensions')
# options.add_argument('--no-sandbox')
# options.add_argument('--disable-infobars')
# options.add_argument('--disable-dev-shm-usage')
# options.add_argument('--disable-browser-side-navigation')
# options.add_argument('--disable-gpu')
# # options.add_argument(r"C:\Users\hecto\AppData\Local\Google\Chrome\User Data\Persona 1")
#
# # options.page_load_strategy = 'none'
#
# # Iniciar el navegador con las opciones configuradas
# driver = webdriver.Chrome(options=options)
#
# # Iniciar sesión en LinkedIn
# driver.execute_script("window.open('https://www.linkedin.com/home')")
# time.sleep(10)
# driver.switch_to.window(driver.window_handles[1])
# time.sleep(12)
#
# username_field = driver.find_element(By.XPATH, '//*[@id="session_key"]').send_keys("arqhector08@gmail.com")
# time.sleep(3)
# password_field = driver.find_element(By.XPATH, '//*[@id="session_password"]').send_keys("alejandro20759364")
# time.sleep(5)
# btn = driver.find_element(By.XPATH, '//*[@id="main-content"]/section[1]/div/div/form/div[2]/button')
# btn.click()
# time.sleep(15)
#
#
# # Leer el archivo Excel y obtener los URLs de la columna "Referencia"
# df = pd.read_excel('linkedin_links.xlsx', sheet_name='Sheet1', usecols=['Enlace'])
#
# # Convertir los URLs en una lista
# url_list = df['Enlace'].tolist()
#
#
#
# # Definir un diccionario que mapee el número de intento al XPath correspondiente
# xpath_dict = {
#     1: "//section[@class='pv-contact-info__contact-type'][2]/ul/li/a",  # Web
#     2: "//*[@class='pv-contact-info__contact-type'][3]/ul/li/span",         # Teléfono
#     3: "//section[@class='pv-contact-info__contact-type'][4]/div/a"     # Email
# }
#
# # Crear un nuevo archivo Excel
# wb = openpyxl.Workbook()
# sheet = wb.active
#
# # Agregar encabezados a las columnas
# sheet.append(["Name", "Web", "Phone", "Email"])
#
# # Definir un contador para realizar el seguimiento del número de contactos procesados
# contact_count = 0
#
# # Recorrer cada URL de la lista
# for url in url_list:
#     max_attempts = 3  # Establecer el número total de intentos
#
#     # Definir una lista para almacenar los datos de contacto de cada URL
#     contact_data = [None, None, None, None]  # [Nombre, Web, Teléfono, Email]
#
#     # Iterar sobre los intentos
#     for attempt in range(1, max_attempts + 1):
#         # try:
#         driver.get(url)
#         time.sleep(40)
#
#         # Extraer el nombre solo en la primera iteración
#         if attempt == 1:
#             try:
#                 name = driver.find_element(By.XPATH, "//div[@class='oSsIKjSDcyduLrRLQBBCYBpUDyeLCY']/span/a/h1")
#                 contact_data[0] = name.text
#             except NoSuchElementException:
#                 print(f"No se encontró el elemento en la URL: {url}")
#                 contact_data[0] = "N/A"
#
#         # Intenta presionar el botón de información
#         try:
#             information_button = WebDriverWait(driver, 30).until(
#                 EC.element_to_be_clickable((By.ID, 'top-card-text-details-contact-info'))
#             )
#             actions = ActionChains(driver)
#             actions.move_to_element(information_button).click().perform()
#
#         except NoSuchElementException:
#             print(f"No se encontró el botón de información para la URL: {url}")
#             continue
#
#         time.sleep(30)
#
#         # Intenta encontrar el elemento correspondiente al intento actual
#         xpath = xpath_dict.get(attempt)
#         if xpath:
#             try:
#                 element = driver.find_element(By.XPATH, xpath)
#                 contact_data[attempt] = element.text if attempt != 1 else element.get_attribute("href")
#             except NoSuchElementException:
#                 print(f"No se encontró el elemento para el intento {attempt} en la URL: {url}")
#                 contact_data[attempt] = "N/A"
#         else:
#             print(f"No se proporcionó un XPath para el intento {attempt}")
#
#     # Imprimir los datos de contacto
#     print("\nDatos de Perfil:")
#     print(f"Nombre: {contact_data[0]}")
#     print(f"Web: {contact_data[1]}")
#     print(f"Teléfono: {contact_data[2]}")
#     print(f"Email: {contact_data[3]}")
#
#     # Guardar los datos de contacto en el archivo Excel
#     sheet.append(contact_data)
#     contact_count += 1
#
#     # Guardar en un archivo Excel cada 10 contactos
#     if contact_count % 10 == 0:
#         wb.save("contact_data.xlsx")
#         print("Datos guardados en el archivo Excel.")
#
# # Cerrar el navegador
# driver.close()
#
# # Guardar el archivo Excel final
# wb.save("contact_data.xlsx")







import pandas as pd
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl
import os
import os
from dotenv import load_dotenv
load_dotenv()

# Configuración de las opciones del navegador
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option('useAutomationExtension', False)
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument('--disable-extensions')
options.add_argument('--no-sandbox')
options.add_argument('--disable-infobars')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--disable-browser-side-navigation')
options.add_argument('--disable-gpu')
# options.add_argument(r"C:\Users\hecto\AppData\Local\Google\Chrome\User Data\Persona 1")

# options.page_load_strategy = 'none'

# Iniciar el navegador con las opciones configuradas
driver = webdriver.Chrome(options=options)

# Iniciar sesión en LinkedIn
driver.execute_script("window.open('https://www.linkedin.com/home')")
time.sleep(10)
driver.switch_to.window(driver.window_handles[1])
time.sleep(12)

username_field = driver.find_element(By.XPATH, '//*[@id="session_key"]').send_keys(os.environ.get('LINKEDIN_USERNAME'))
time.sleep(3)
password_field = driver.find_element(By.XPATH, '//*[@id="session_password"]').send_keys(os.environ.get('LINKEDIN_PASSWORD'))
time.sleep(5)
btn = driver.find_element(By.XPATH, '//*[@id="main-content"]/section[1]/div/div/form/div[2]/button')
btn.click()
time.sleep(15)


# Leer el archivo Excel y obtener los URLs de la columna "Referencia"
df = pd.read_excel('linkedin_links.xlsx', sheet_name='Sheet1', usecols=['Enlace'])

# Convertir los URLs en una lista
url_list = df['Enlace'].tolist()



# Definir un diccionario que mapee el número de intento al XPath correspondiente
xpath_dict = {
    1: "//section[@class='pv-contact-info__contact-type'][2]/ul/li/a",  # Web
    2: "//*[@class='pv-contact-info__contact-type'][3]/ul/li/span",         # Teléfono
    3: "//section[@class='pv-contact-info__contact-type'][4]/div/a"     # Email
}

# Crear un nuevo archivo Excel
wb = openpyxl.Workbook()
sheet = wb.active

# Agregar encabezados a las columnas
sheet.append(["Name", "Web", "Phone", "Email"])

# Definir un contador para realizar el seguimiento del número de contactos procesados
contact_count = 0

# Recorrer cada URL de la lista
for url in url_list:
    try:
        # Intentar acceder a la página y extraer los datos
        max_attempts = 3  # Establecer el número total de intentos

        # Definir variables para almacenar los datos de contacto
        value_name = ""
        value_web = ""
        value_phone = ""
        value_email = ""

        for attempt in range(1, max_attempts + 1):
            # try:
            driver.get(url)
            time.sleep(40)

            # Extraer el nombre solo en la primera iteración
            if attempt == 1:
                try:
                    name = driver.find_element(By.XPATH, "//div[@class='oSsIKjSDcyduLrRLQBBCYBpUDyeLCY']/span/a/h1")
                    value_name = name.text
                except NoSuchElementException:
                    print(f"No se encontró el elemento en la URL: {url}")
                    value_name = "N/A"

            # Intenta presionar el botón de información
            try:
                information_button = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((By.ID, 'top-card-text-details-contact-info'))
                )
                actions = ActionChains(driver)
                actions.move_to_element(information_button).click().perform()

            except NoSuchElementException:
                print(f"No se encontró el botón de información para la URL: {url}")
                continue

            time.sleep(30)

            # Intenta encontrar el elemento correspondiente al intento actual
            xpath = xpath_dict.get(attempt)
            if xpath:
                try:
                    element = driver.find_element(By.XPATH, xpath)
                    if attempt == 1:
                        value_web = element.get_attribute("href")
                    elif attempt == 2:
                        value_phone = element.text
                    elif attempt == 3:
                        value_email = element.text
                except NoSuchElementException:
                    print(f"No se encontró el elemento para el intento {attempt} en la URL: {url}")
            else:
                print(f"No se proporcionó un XPath para el intento {attempt}")

        # Imprimir los datos de contacto
        print("\nDatos de Perfil:")
        print(f"Nombre: {value_name}")
        print(f"Web: {value_web}")
        print(f"Teléfono: {value_phone}")
        print(f"Email: {value_email}")

        # Guardar los datos de contacto en el archivo Excel
        sheet.append([value_name, value_web, value_phone, value_email])
        contact_count += 1

        # Guardar en un archivo Excel cada 10 contactos
        if contact_count % 5 == 0:
            wb.save("contact_data.xlsx")
            print("Datos guardados en el archivo Excel.")

    except Exception as e:
        print(f"Se produjo un error al procesar la URL {url}: {str(e)}")
        continue  # Continuar con el siguiente enlace si se produce un error

# Cerrar el navegador
driver.close()

# Guardar el archivo Excel final
wb.save("contact_data.xlsx")